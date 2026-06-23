#!/usr/bin/env python3
"""
批签发证明拆分工具 — Web 服务
启动后自动打开浏览器，通过 http://127.0.0.1:5050 访问。
"""
from __future__ import annotations

import json
import os
import sys
import tempfile
import uuid
import zipfile
from datetime import datetime
from pathlib import Path

from flask import Flask, Response, jsonify, request, send_file, send_from_directory

# 将 tools/ 加入路径，复用核心逻辑
sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "tools"))
import batch_cert_splitter as core

if getattr(sys, "frozen", False):
    _STATIC_DIR = Path(sys._MEIPASS) / "static"  # type: ignore[attr-defined]
else:
    _STATIC_DIR = Path(__file__).parent / "static"

app = Flask(__name__, static_folder=str(_STATIC_DIR), static_url_path="")

# 上传文件临时目录
WORK_DIR = Path(tempfile.gettempdir()) / "cert_splitter_web"
WORK_DIR.mkdir(parents=True, exist_ok=True)

# ── 静态页面 ──────────────────────────────────────────────────
@app.route("/")
def index():
    return send_from_directory("static", "index.html")


# ── 检测接口（SSE 流式返回进度） ──────────────────────────────
@app.route("/api/detect", methods=["POST"])
def api_detect():
    if "pdf" not in request.files:
        return jsonify({"error": "请选择 PDF 文件"}), 400

    f = request.files["pdf"]
    if not f.filename.lower().endswith(".pdf"):
        return jsonify({"error": "只支持 PDF 文件"}), 400

    no_qr = request.form.get("no_qr") == "1"
    dpi   = int(request.form.get("dpi", 200))

    # 保存上传的 PDF
    file_id  = uuid.uuid4().hex
    pdf_path = WORK_DIR / f"{file_id}.pdf"
    f.save(str(pdf_path))
    pdf_name = f.filename

    def generate():
        import fitz

        try:
            doc         = fitz.open(str(pdf_path))
            total_pages = len(doc)
        except Exception as e:
            yield _sse({"type": "error", "message": f"PDF 打开失败：{e}"})
            return

        yield _sse({"type": "start", "total": total_pages, "file_id": file_id, "pdf_name": pdf_name})

        # 判断是否需要 OCR
        try:
            use_ocr = len(doc[0].get_text().strip()) < 10
        except Exception:
            use_ocr = True

        # 判断是否支持 QR
        use_qr = not no_qr
        if use_qr:
            try:
                import cv2  # noqa: F401
            except ImportError:
                use_qr = False
                yield _sse({"type": "warn", "message": "opencv-python 未安装，已切换为纯 OCR 模式"})

        if use_qr:
            yield _sse({"type": "info", "message": "已启用二维码模式（联网获取国家药监局数据）"})
        else:
            engine = "macOS Vision" if sys.platform == "darwin" else "Tesseract"
            yield _sse({"type": "info", "message": f"OCR 模式（{engine}，DPI={dpi}）"})

        detections: list[dict] = []

        for i in range(total_pages):
            page = doc[i]

            # 优先二维码
            if use_qr:
                try:
                    qr = core._decode_qr_page(page)
                    if qr and "nmpa.gov.cn" in qr:
                        fields = core._fetch_nmpa_cert(qr)
                        if fields and fields.get("batch_no"):
                            detections.append({"page": i + 1, "conf": "高", "source": "qr", "fields": fields})
                            yield _sse({
                                "type": "page", "page": i + 1, "total": total_pages,
                                "source": "qr", "batch_no": fields.get("batch_no", ""),
                                "conf": "高",
                            })
                            continue
                except Exception:
                    pass

            # OCR 兜底
            try:
                text = core._ocr_page(page, dpi=dpi) if use_ocr else page.get_text()
            except Exception as e:
                yield _sse({"type": "error", "message": f"第 {i+1} 页 OCR 失败：{e}"})
                return

            conf, fields = core._score_page(text)
            yield _sse({
                "type": "page", "page": i + 1, "total": total_pages,
                "source": "ocr" if conf else "none",
                "batch_no": fields.get("batch_no", "") if conf else "",
                "conf": conf or "",
            })
            if conf:
                detections.append({"page": i + 1, "conf": conf, "source": "ocr", "fields": fields})

        # 构建索引
        records, warnings = core._build_detect_index(detections, total_pages)

        yield _sse({
            "type":        "done",
            "file_id":     file_id,
            "pdf_name":    pdf_name,
            "total_pages": total_pages,
            "records":     records,
            "warnings":    warnings,
        })

    return Response(
        generate(),
        mimetype="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


# ── 拆分接口 ──────────────────────────────────────────────────
@app.route("/api/split", methods=["POST"])
def api_split():
    import fitz

    data     = request.json or {}
    file_id  = data.get("file_id", "")
    pdf_name = data.get("pdf_name", "output.pdf")
    records  = data.get("records", [])
    force    = bool(data.get("force", False))

    if not file_id or not records:
        return jsonify({"error": "参数不完整"}), 400

    pdf_path = WORK_DIR / f"{file_id}.pdf"
    if not pdf_path.exists():
        return jsonify({"error": "PDF 已过期，请重新上传"}), 404

    # 补全字段类型
    for r in records:
        r["start_page"] = int(r.get("start_page") or 0)
        r["end_page"]   = int(r.get("end_page") or 0)
        r.setdefault("batch_no",     "")
        r.setdefault("vaccine_name", None)
        r.setdefault("manufacturer", None)
        r.setdefault("cert_no",      None)
        r.setdefault("notes",        None)

    try:
        src         = fitz.open(str(pdf_path))
        total_pages = len(src)
    except Exception as e:
        return jsonify({"error": f"PDF 打开失败：{e}"}), 500

    # 质控
    errors = core._validate(records, total_pages)
    hard   = [e for e in errors if not e.startswith("警告")]
    if hard:
        return jsonify({"error": "\n".join(hard)}), 422

    # 重复拆分检查（写文件前）
    if not force:
        try:
            core.init_db()
            with core.get_conn() as conn:
                cur = conn.execute(
                    "INSERT OR IGNORE INTO batch_files "
                    "(original_filename, original_path, total_pages, import_time, status) "
                    "VALUES (?,?,?,?,'split')",
                    (pdf_name, str(pdf_path), total_pages, datetime.now().isoformat(timespec="seconds")),
                )
                db_file_id_check = cur.lastrowid or conn.execute(
                    "SELECT id FROM batch_files WHERE original_path=?", (str(pdf_path),)
                ).fetchone()["id"]
                prev_count = conn.execute(
                    "SELECT COUNT(*) FROM cert_index WHERE batch_file_id=?", (db_file_id_check,)
                ).fetchone()[0]
                if prev_count > 0:
                    return jsonify({"already_split": True, "prev_count": prev_count})
        except Exception:
            pass

    # 输出目录
    out_dir = WORK_DIR / file_id / "output"
    out_dir.mkdir(parents=True, exist_ok=True)

    results = []
    for r in records:
        s, e       = r["start_page"], r["end_page"]
        page_count = e - s + 1

        stem    = r["batch_no"] or f"第{s}-{e}页"
        if r.get("cert_no"):
            stem = f"{stem}_{r['cert_no']}"
        # 过滤文件名非法字符
        safe_stem = "".join(c if c not in r'\/:*?"<>|' else "_" for c in stem)
        out_pdf   = out_dir / f"{safe_stem}.pdf"

        review     = "ok"
        notes_list = [r["notes"]] if r.get("notes") else []
        lo, hi     = core.NORMAL_PAGE_RANGE
        if not (lo <= page_count <= hi):
            review = "abnormal"
            notes_list.append(f"页数={page_count}，超出正常范围，需人工复核")

        writer = fitz.open()
        writer.insert_pdf(src, from_page=s - 1, to_page=e - 1)
        writer.save(str(out_pdf))
        writer.close()

        results.append({
            **r,
            "page_count":    page_count,
            "output_pdf":    str(out_pdf),
            "review_status": review,
            "notes":         "；".join(notes_list) or None,
        })

    # 数据库入库
    try:
        core.init_db()
        now = datetime.now().isoformat(timespec="seconds")
        with core.get_conn() as conn:
            cur = conn.execute(
                "INSERT OR IGNORE INTO batch_files "
                "(original_filename, original_path, total_pages, import_time, status) "
                "VALUES (?,?,?,?,'split')",
                (pdf_name, str(pdf_path), total_pages, now),
            )
            db_file_id = cur.lastrowid or conn.execute(
                "SELECT id FROM batch_files WHERE original_path=?", (str(pdf_path),)
            ).fetchone()["id"]

            # force=True 时清除旧记录后重新写入
            conn.execute("DELETE FROM cert_index WHERE batch_file_id=?", (db_file_id,))

            split_time = datetime.now().isoformat(timespec="seconds")
            for r in results:
                conn.execute(
                    "INSERT INTO cert_index "
                    "(batch_file_id, batch_no, vaccine_name, manufacturer, cert_no,"
                    " start_page, end_page, page_count, output_pdf, review_status, split_time, notes)"
                    " VALUES (?,?,?,?,?,?,?,?,?,?,?,?)",
                    (db_file_id, r["batch_no"], r["vaccine_name"], r["manufacturer"],
                     r["cert_no"], r["start_page"], r["end_page"], r["page_count"],
                     r["output_pdf"], r["review_status"], split_time, r["notes"]),
                )
    except Exception:
        pass  # 入库失败不影响文件下载

    # 打包 ZIP
    stem_name = Path(pdf_name).stem
    zip_path  = WORK_DIR / file_id / f"{stem_name}_拆分结果.zip"
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
        for r in results:
            zf.write(r["output_pdf"], Path(r["output_pdf"]).name)

    abnormal = [r for r in results if r["review_status"] == "abnormal"]
    return jsonify({
        "ok":          True,
        "count":       len(results),
        "abnormal":    len(abnormal),
        "download_id": file_id,
        "zip_name":    zip_path.name,
        "results":     [
            {"batch_no": r["batch_no"], "pages": f"{r['start_page']}-{r['end_page']}",
             "page_count": r["page_count"], "review_status": r["review_status"]}
            for r in results
        ],
    })


# ── 下载 ZIP ──────────────────────────────────────────────────
@app.route("/api/download/<file_id>")
def api_download(file_id: str):
    # 安全检查：file_id 只含十六进制字符
    if not all(c in "0123456789abcdef" for c in file_id):
        return jsonify({"error": "非法请求"}), 400

    out_dir = WORK_DIR / file_id
    zips    = list(out_dir.glob("*.zip"))
    if not zips:
        return jsonify({"error": "文件不存在"}), 404

    return send_file(str(zips[0]), as_attachment=True, download_name=zips[0].name)


# ── 导出 Excel 索引 ───────────────────────────────────────────
@app.route("/api/export_excel/<file_id>")
def api_export_excel(file_id: str):
    import io
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    if not all(c in "0123456789abcdef" for c in file_id):
        return jsonify({"error": "非法请求"}), 400

    pdf_path = WORK_DIR / f"{file_id}.pdf"

    try:
        core.init_db()
        with core.get_conn() as conn:
            bf = conn.execute(
                "SELECT id, original_filename FROM batch_files WHERE original_path=?",
                (str(pdf_path),),
            ).fetchone()
            if not bf:
                return jsonify({"error": "记录不存在，请先完成拆分"}), 404

            rows = conn.execute(
                "SELECT batch_no, vaccine_name, manufacturer, cert_no,"
                " start_page, end_page, page_count, review_status, split_time, notes"
                " FROM cert_index WHERE batch_file_id=? ORDER BY start_page",
                (bf["id"],),
            ).fetchall()
    except Exception as e:
        return jsonify({"error": str(e)}), 500

    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title = "批签发索引"

    hdr_font  = Font(bold=True, color="FFFFFF", size=11)
    hdr_fill  = PatternFill("solid", fgColor="1A6FC4")
    title_font = Font(bold=True, size=13)
    ok_fill   = PatternFill("solid", fgColor="D6F0E0")
    warn_fill = PatternFill("solid", fgColor="FFF8D6")
    thin_side = Side(style="thin", color="DDDDDD")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    center    = Alignment(horizontal="center", vertical="center")
    left_mid  = Alignment(horizontal="left",   vertical="center")

    # 标题行
    pdf_name   = bf["original_filename"]
    split_date = (rows[0]["split_time"] if rows else "").split("T")[0]
    ws.merge_cells("A1:J1")
    ws["A1"] = f"批签发证明索引 — {pdf_name}（拆分时间：{split_date}）"
    ws["A1"].font      = title_font
    ws["A1"].alignment = left_mid
    ws.row_dimensions[1].height = 28

    # 表头行
    headers = ["序号", "批号", "疫苗名称", "生产企业", "批签发号",
               "起始页", "结束页", "页数", "状态", "备注"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=col, value=h)
        c.font = hdr_font; c.fill = hdr_fill
        c.alignment = center; c.border = thin_border
    ws.row_dimensions[2].height = 22

    # 数据行
    CENTER_COLS = {1, 6, 7, 8, 9}
    for i, r in enumerate(rows, 1):
        is_ok  = r["review_status"] == "ok"
        fill   = ok_fill if is_ok else warn_fill
        values = [
            i,
            r["batch_no"],
            r["vaccine_name"] or "",
            r["manufacturer"] or "",
            r["cert_no"]      or "",
            r["start_page"],
            r["end_page"],
            r["page_count"],
            "正常" if is_ok else "待复核",
            r["notes"] or "",
        ]
        rn = i + 2
        for col, val in enumerate(values, 1):
            c = ws.cell(row=rn, column=col, value=val)
            c.fill = fill; c.border = thin_border
            c.alignment = center if col in CENTER_COLS else left_mid
        ws.row_dimensions[rn].height = 18

    # 列宽
    for col, w in enumerate([6, 22, 24, 22, 22, 8, 8, 6, 8, 30], 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.freeze_panes = "A3"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    excel_name = f"{Path(pdf_name).stem}_批签发索引.xlsx"
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=excel_name,
    )


# ── 历史记录接口 ──────────────────────────────────────────────
@app.route("/api/records")
def api_records():
    try:
        core.init_db()
        with core.get_conn() as conn:
            rows = conn.execute(
                "SELECT c.batch_no, c.vaccine_name, c.cert_no, c.start_page, c.end_page,"
                "       c.page_count, c.review_status, c.split_time, b.original_filename"
                " FROM cert_index c JOIN batch_files b ON b.id=c.batch_file_id"
                " ORDER BY c.id DESC LIMIT 200"
            ).fetchall()
        return jsonify([dict(r) for r in rows])
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ── 工具函数 ──────────────────────────────────────────────────
def _sse(data: dict) -> str:
    return f"data: {json.dumps(data, ensure_ascii=False)}\n\n"


# ── 启动入口 ──────────────────────────────────────────────────
def main():
    import threading
    import webbrowser

    port = 5050
    url  = f"http://127.0.0.1:{port}"

    def open_browser():
        import time
        time.sleep(1.2)
        webbrowser.open(url)

    threading.Thread(target=open_browser, daemon=True).start()
    print(f"\n批签发证明拆分工具已启动: {url}")
    print("关闭此窗口即可退出程序\n")
    app.run(host="127.0.0.1", port=port, debug=False, threaded=True)


if __name__ == "__main__":
    main()
