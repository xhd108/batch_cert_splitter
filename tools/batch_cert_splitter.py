#!/usr/bin/env python3
"""
批签发大 PDF 拆分工具 v2

子命令：
  init      初始化 SQLite 数据库
  template  生成人工填写用的 Excel 索引模板（v1 流程）
  detect    自动识别疑似首页，生成复核表（v2 流程）
  split     按索引/复核表拆分 PDF，写入数据库，生成拆分报告
  list      查看已入库的单证记录

v2 用法示例：
  python3 tools/batch_cert_splitter.py detect --pdf 批签发.pdf
  # 用 Excel 打开生成的复核表，核对/修正后保存
  python3 tools/batch_cert_splitter.py split --pdf 批签发.pdf --index 批签发_复核表.xlsx

v1 用法示例（人工填写）：
  python3 tools/batch_cert_splitter.py template
  python3 tools/batch_cert_splitter.py split --pdf 批签发.pdf --index 批签发索引模板.xlsx
"""
from __future__ import annotations

import argparse
import logging
import re
import sqlite3
import sys
from datetime import datetime
from pathlib import Path

import fitz  # PyMuPDF
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

# ── 数据根目录（兼容 PyInstaller 打包环境） ──────────────────────
import os as _os
_env_data = _os.environ.get("CERT_SPLITTER_DATA")
if _env_data:
    DATA_ROOT = Path(_env_data)
elif getattr(sys, "frozen", False):
    DATA_ROOT = Path.home() / ".cert_splitter"
else:
    DATA_ROOT = Path(__file__).resolve().parent.parent

DB_PATH    = DATA_ROOT / "db" / "cert.db"
OUTPUT_DIR = DATA_ROOT / "output"
LOG_DIR    = DATA_ROOT / "logs"

# 页数不在此区间则标记为待复核（4页证明属正常，扩展上限）
NORMAL_PAGE_RANGE = (2, 4)

# ── OCR：macOS Vision 框架 ────────────────────────────────────
def _dewarp_image(img: "np.ndarray") -> "np.ndarray | None":
    """
    检测文档边界四边形，应用透视变换矫正拍照倾斜/弯曲。
    成功返回矫正后 RGB 图像，无法检测四边形时返回 None。
    适用于手机扫描件（CamScanner 等）透视畸变校正。
    """
    try:
        import cv2
        import numpy as np
    except ImportError:
        return None

    gray = cv2.cvtColor(img, cv2.COLOR_RGB2GRAY) if img.ndim == 3 else img
    edges = cv2.Canny(cv2.GaussianBlur(gray, (5, 5), 0), 30, 100)
    edges = cv2.dilate(edges, np.ones((3, 3), np.uint8), iterations=2)

    contours, _ = cv2.findContours(edges, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    h, w = gray.shape
    best, best_area = None, 0
    for cnt in contours:
        area = cv2.contourArea(cnt)
        if area < w * h * 0.1:
            continue
        approx = cv2.approxPolyDP(cnt, 0.02 * cv2.arcLength(cnt, True), True)
        if len(approx) == 4 and area > best_area:
            best_area, best = area, approx

    if best is None:
        return None

    pts = best.reshape(4, 2).astype(np.float32)
    s = pts.sum(axis=1)
    d = np.diff(pts, axis=1).ravel()
    tl, br = pts[s.argmin()], pts[s.argmax()]
    tr, bl = pts[d.argmin()], pts[d.argmax()]
    src = np.array([tl, tr, br, bl], dtype=np.float32)

    w_out = int(max(np.linalg.norm(br - bl), np.linalg.norm(tr - tl)))
    h_out = int(max(np.linalg.norm(tr - br), np.linalg.norm(tl - bl)))
    if w_out < 100 or h_out < 100:
        return None

    dst = np.array([[0, 0], [w_out-1, 0], [w_out-1, h_out-1], [0, h_out-1]], dtype=np.float32)
    M = cv2.getPerspectiveTransform(src, dst)
    return cv2.warpPerspective(img, M, (w_out, h_out))


def _ocr_page_vision(page: "fitz.Page", dpi: int = 200,
                     img: "np.ndarray | None" = None) -> str:
    """
    用 macOS Vision 框架对 PDF 单页进行 OCR，返回识别文字字符串。
    img：若提供预渲染（如已矫正的）图像则直接使用，否则从 page 渲染。
    """
    try:
        import Vision
        import Quartz
    except ImportError:
        raise RuntimeError(
            "macOS Vision 框架未安装。请运行：\n"
            "  pip3 install pyobjc-framework-Vision pyobjc-framework-Quartz"
        )

    import tempfile, os

    with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as f:
        tmp_path = f.name
    try:
        if img is not None:
            from PIL import Image as _PILImage
            _PILImage.fromarray(img).save(tmp_path)
        else:
            scale = dpi / 72
            pix = page.get_pixmap(matrix=fitz.Matrix(scale, scale), colorspace=fitz.csGRAY)
            pix.save(tmp_path)

        url = Quartz.NSURL.fileURLWithPath_(tmp_path)
        req = Vision.VNRecognizeTextRequest.alloc().init()
        req.setRecognitionLanguages_(["zh-Hans", "zh-Hant", "en-US"])
        req.setRecognitionLevel_(Vision.VNRequestTextRecognitionLevelAccurate)
        req.setUsesLanguageCorrection_(False)

        handler = Vision.VNImageRequestHandler.alloc().initWithURL_options_(url, None)
        handler.performRequests_error_([req], None)

        lines = [
            obs.topCandidates_(1)[0].string()
            for obs in (req.results() or [])
            if obs.topCandidates_(1)
        ]
        return "\n".join(lines)
    finally:
        os.unlink(tmp_path)


# ── OCR：Linux / 麒麟 Tesseract ──────────────────────────────
def _ocr_page_tesseract(page: "fitz.Page", dpi: int = 200,
                        img: "np.ndarray | None" = None) -> str:
    """
    用 Tesseract OCR 对 PDF 单页进行识别，返回文字字符串。
    img：若提供预渲染（如已矫正的）图像则直接使用，否则从 page 渲染。
    """
    try:
        import pytesseract
    except ImportError:
        raise RuntimeError(
            "pytesseract 未安装。请运行：\n"
            "  sudo apt install tesseract-ocr tesseract-ocr-chi-sim tesseract-ocr-chi-tra\n"
            "  pip3 install pytesseract"
        )

    import tempfile, os

    with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as f:
        tmp_path = f.name
    try:
        if img is not None:
            from PIL import Image as _PILImage
            _PILImage.fromarray(img).save(tmp_path)
        else:
            scale = dpi / 72
            pix = page.get_pixmap(matrix=fitz.Matrix(scale, scale), colorspace=fitz.csGRAY)
            pix.save(tmp_path)
        return pytesseract.image_to_string(
            tmp_path,
            lang="chi_sim+chi_tra+eng",
            config="--psm 6",
        )
    finally:
        os.unlink(tmp_path)


def _ocr_page(page: "fitz.Page", dpi: int = 200,
              img: "np.ndarray | None" = None) -> str:
    """自动选择 OCR 引擎：macOS 用 Vision，其他系统用 Tesseract。"""
    if sys.platform == "darwin":
        return _ocr_page_vision(page, dpi, img=img)
    else:
        return _ocr_page_tesseract(page, dpi, img=img)


# ── 二维码 + 国家药监局 API ───────────────────────────────────
def _decode_qr_page(page: "fitz.Page", scale: float = 3.0) -> str | None:
    """
    从 PDF 页面解码二维码，返回解码内容字符串，失败返回 None。
    优先使用 zxing-cpp（支持 ECI 编码，国家药监局 QR 码必须），
    cv2 作为兜底（全页 + 四角裁剪 + Otsu 二值化）。
    """
    try:
        import numpy as np
    except ImportError:
        return None

    mat = fitz.Matrix(scale, scale)
    pix = page.get_pixmap(matrix=mat)
    img = np.frombuffer(pix.samples, dtype=np.uint8).reshape(pix.height, pix.width, pix.n)

    # ── 首选：zxing-cpp（完整支持 ECI/UTF-8 编码的 QR 码） ───────
    try:
        import zxingcpp
        results = zxingcpp.read_barcodes(img)
        for r in results:
            if r.format == zxingcpp.BarcodeFormat.QRCode and r.text:
                return r.text
    except Exception:
        pass

    # ── 兜底：cv2 ─────────────────────────────────────────────────
    # 策略1：全页（原图/灰度/锐化/Otsu二值化）
    # 策略2：四角裁剪（1/5边长）+ Otsu 二值化
    # 扫描件 QR 码偏小且可能倾斜，裁剪 + 二值化显著提升检出率
    try:
        import cv2
        if pix.n == 4:
            img_rgb = cv2.cvtColor(img, cv2.COLOR_RGBA2RGB)
        else:
            img_rgb = img
        gray = cv2.cvtColor(img_rgb, cv2.COLOR_RGB2GRAY)
        detector = cv2.QRCodeDetector()
        kernel = np.array([[-1, -1, -1], [-1, 9, -1], [-1, -1, -1]])
        _, thresh = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)

        # 策略1：全页
        for src in [img_rgb, gray, cv2.filter2D(gray, -1, kernel), thresh]:
            data, _, _ = detector.detectAndDecode(src)
            if data:
                return data

        # 策略2：四角高分辨率单独渲染（scale=3 时 QR 码约60px 宽，偏小；
        #         直接对角落区域用 scale=8 渲染，约 160px 宽，检出率显著提升）
        pw, ph = page.rect.width, page.rect.height
        hi_scale = 8.0
        hi_mat = fitz.Matrix(hi_scale, hi_scale)
        corner_fracs = [
            (0, 0.2, 0, 0.2),      # 左上
            (0, 0.2, 0.8, 1.0),    # 右上
            (0.8, 1.0, 0, 0.2),    # 左下
            (0.8, 1.0, 0.8, 1.0),  # 右下
        ]
        for y0f, y1f, x0f, x1f in corner_fracs:
            clip = fitz.Rect(x0f * pw, y0f * ph, x1f * pw, y1f * ph)
            hi_pix = page.get_pixmap(matrix=hi_mat, clip=clip)
            hi_img = np.frombuffer(hi_pix.samples, dtype=np.uint8).reshape(
                hi_pix.height, hi_pix.width, hi_pix.n
            )
            hi_gray = cv2.cvtColor(
                cv2.cvtColor(hi_img, cv2.COLOR_RGBA2RGB) if hi_pix.n == 4 else hi_img,
                cv2.COLOR_RGB2GRAY,
            )
            _, hi_thresh = cv2.threshold(hi_gray, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
            for src in [hi_gray, hi_thresh]:
                data, _, _ = detector.detectAndDecode(src)
                if data:
                    return data

        # 策略3：透视矫正后重试（处理拍照倾斜导致 QR 码畸变的情况）
        dewarped = _dewarp_image(img_rgb)
        if dewarped is not None:
            import zxingcpp as _zxing
            for r in _zxing.read_barcodes(dewarped):
                if r.format == _zxing.BarcodeFormat.QRCode and r.text:
                    return r.text
            dw_gray = cv2.cvtColor(dewarped, cv2.COLOR_RGB2GRAY)
            _, dw_thresh = cv2.threshold(dw_gray, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
            dw_h, dw_w = dw_gray.shape
            dw_r = 5
            for crop_g in [
                dw_gray, dw_thresh,
                dw_gray[: dw_h // dw_r, : dw_w // dw_r],
                dw_thresh[: dw_h // dw_r, : dw_w // dw_r],
            ]:
                data, _, _ = detector.detectAndDecode(crop_g)
                if data:
                    return data
    except Exception:
        pass

    return None


def _fetch_nmpa_cert(qr_url: str, timeout: int = 8) -> dict | None:
    """
    调用国家药监局验证 API，解析返回的 HTML 中嵌入的 JSON，
    提取批号、批签发号、疫苗名称、生产企业。
    返回 None 表示网络失败或解析失败。
    """
    import urllib.request, json as _json

    try:
        req = urllib.request.Request(
            qr_url, headers={"User-Agent": "Mozilla/5.0"}
        )
        with urllib.request.urlopen(req, timeout=timeout) as r:
            html = r.read().decode("utf-8", errors="replace")
    except Exception:
        return None

    def _parse_surface(surface_list: list, cert_no_fallback: str | None) -> dict:
        surface = {s["name"]: s["value"] for s in surface_list}

        def _get(*keys):
            for k in keys:
                v = surface.get(k, "").strip()
                if v:
                    return v
            return None

        return {
            "batch_no":     _get("批号(Lot No.)", "批号"),
            "cert_no":      _get("Certificate No.") or cert_no_fallback or None,
            "vaccine_name": _get("产品名称(Generic Name)", "产品名称"),
            "manufacturer": _get("生产企业(Manufacturer)", "生产企业"),
        }

    # ── 格式1：sels.nmpa.gov.cn — var Certificate = '{...}'; ──────
    m = re.search(r"var Certificate = '(\{.*?\})';", html, re.DOTALL)
    if m:
        try:
            data = _json.loads(m.group(1))
            items = data.get("data", {}).get("dataList", [])
            if items:
                cert = items[0]
                cn_fallback = cert.get("certificateNumber", "").strip() or None
                return _parse_surface(cert.get("surface", []), cn_fallback)
        except Exception:
            pass

    # ── 格式2：zhjg.nmpa.gov.cn — surface 数组后紧跟 certificateName ─
    m2 = re.search(r'"surface":\s*\[(.+?)\],\s*"certificateName"', html, re.DOTALL)
    if m2:
        try:
            surface_list = _json.loads('[' + m2.group(1) + ']')
            cn_m = re.search(r'"certificateNumber":\s*"([^"]+)"', html)
            cn_fallback = cn_m.group(1).strip() if cn_m else None
            return _parse_surface(surface_list, cn_fallback)
        except Exception:
            pass

    return None


# ── 自动识别：首页特征 ────────────────────────────────────────
# 高置信关键词分两类：
#   EXACT  — 直接子串匹配（这些词不会出现在正文句子中）
#   LINE   — 必须在行首出现（避免被"已获得批签发合格证书"等语句误触发）
_HIGH_KW_EXACT = [
    "批签发证明",
    "Certificate for the Release of Biological",
    "Lot Release Certificate",
]
_HIGH_KW_LINE = re.compile(
    r'^\s*(?:生物制品批签发|批签发合格证|疫苗批签发)',
    re.MULTILINE,
)

# 批号：支持五种版式
#   双行版：Lot No. 后换行取批号（国家批签发证明标准版式，含 Lot:No. OCR变形）
#   逆序版：批号值出现在 Lot No./批号 标签之前（OCR按列读序时）
#   双栏版：批号出现在"收检编号"之前的独立行（部分厂家版式）
#   产品批号版：产品批号\n<值>（企业报告格式）
#   单行版：批号：202512057
# 括号后缀统一纳入捕获：202506022（1-2）/ 202506022（-1，-2，-3）/ 202506022（1～2）
_BATCH_SUFFIX = r'(?:[（(][^）)\n]{1,30}[）)])?'   # 可选括号后缀

_RE_BATCH_NO = re.compile(
    r'(?:'
    # 双行版/产品批号版：Lot No. 与批号之间可有 0-8 行 OCR 噪声行（双栏版式混读），
    # 批号必须：含数字 + 总长度≥6（排除 Dosage/注射剂 等短字段名）
    r'(?:Lot[:\s]*No|产品批号)[.：: ]*\s*\n(?:(?!.*?Lot)[^\n]{0,30}\n){0,8}\s*([A-Za-z]{0,3}[0-9][A-Za-z0-9]{4,24}' + _BATCH_SUFFIX + r')'
    r'|([0-9][A-Za-z0-9]{4,11}' + _BATCH_SUFFIX + r')\s*\n\s*Lot[:\s]*No'                    # 逆序版（值在Lot No.前；≤12字符防收检编号误捕获）
    r'|([A-Za-z0-9]{5,20}' + _BATCH_SUFFIX + r')\s*\n\s*收检编号'                            # 双栏版
    r'|批\s*号[：:]\s*([A-Za-z0-9]{3,25}' + _BATCH_SUFFIX + r')'                             # 单行版
    r')',
    re.IGNORECASE,
)

# 批签发号：两种格式均完整保留前缀
#   国际格式：LRA20260923 / LRH20250448 / LRN20260012 / LRG20250220（LR+任意字母+数字）
#   中文格式：批签中检20260923 / 批签鄂检20250448 / 批签甘检20250220（完整字符串）
_RE_CERT_NO_INTL = re.compile(r'LR[A-Z]\d{6,}', re.IGNORECASE)   # 优先匹配国际格式
_RE_CERT_NO_CN   = re.compile(r'批签[^\d\n]{0,8}\d{6,}')          # 回退匹配中文格式

# 疫苗名称：Generic Name 下一行（扫描件标准版式）或含"疫苗"的短行
_RE_VACCINE_GENERIC = re.compile(r'Generic\s*Name\s*\n\s*([^\n]{3,35})')
_RE_VACCINE_LINE    = re.compile(r'^(.{3,35}(?:疫苗|vaccine)[^，。；\n]{0,15})$',
                                  re.MULTILINE | re.IGNORECASE)

# 生产企业：Manufacturer 下一行（扫描件标准版式）或内联版
_RE_MANUFACTURER_NEXT = re.compile(r'Manufacturer[：:\s]*\n\s*([^\n]{2,30})')
_RE_MANUFACTURER      = re.compile(r'生产企业[：:]\s*([^\n]{2,30})')

# ── 日志 ──────────────────────────────────────────────────────
def _setup_logger() -> logging.Logger:
    LOG_DIR.mkdir(parents=True, exist_ok=True)
    log_file = LOG_DIR / f"splitter_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"
    logger = logging.getLogger("cert_splitter")
    logger.setLevel(logging.DEBUG)
    fmt = logging.Formatter("%(asctime)s [%(levelname)s] %(message)s", "%H:%M:%S")
    fh = logging.FileHandler(log_file, encoding="utf-8")
    fh.setLevel(logging.DEBUG)
    fh.setFormatter(fmt)
    ch = logging.StreamHandler(sys.stdout)
    ch.setLevel(logging.INFO)
    ch.setFormatter(fmt)
    logger.addHandler(fh)
    logger.addHandler(ch)
    return logger

log = _setup_logger()


# ── 数据库 ────────────────────────────────────────────────────
def get_conn() -> sqlite3.Connection:
    DB_PATH.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    return conn


def init_db() -> None:
    with get_conn() as conn:
        conn.executescript("""
        CREATE TABLE IF NOT EXISTS batch_files (
            id               INTEGER PRIMARY KEY AUTOINCREMENT,
            original_filename TEXT    NOT NULL,
            original_path    TEXT    NOT NULL UNIQUE,
            total_pages      INTEGER,
            import_time      TEXT,
            status           TEXT    DEFAULT 'pending'
        );

        CREATE TABLE IF NOT EXISTS cert_index (
            id               INTEGER PRIMARY KEY AUTOINCREMENT,
            batch_file_id    INTEGER NOT NULL REFERENCES batch_files(id),
            batch_no         TEXT    NOT NULL,
            vaccine_name     TEXT,
            manufacturer     TEXT,
            cert_no          TEXT,
            start_page       INTEGER NOT NULL,
            end_page         INTEGER NOT NULL,
            page_count       INTEGER,
            output_pdf       TEXT,
            review_status    TEXT    DEFAULT 'pending',
            split_time       TEXT,
            notes            TEXT
        );

        CREATE INDEX IF NOT EXISTS idx_cert_batch_no ON cert_index(batch_no);
        CREATE INDEX IF NOT EXISTS idx_cert_file     ON cert_index(batch_file_id);
        """)
    log.info(f"数据库已就绪: {DB_PATH}")


# ── Excel 模板 ────────────────────────────────────────────────
TEMPLATE_COLS = [
    ("批号",     "batch_no",      "必填，与疫苗标签批号一致"),
    ("疫苗名称", "vaccine_name",  "如：冻干人用狂犬病疫苗（Vero细胞）"),
    ("生产企业", "manufacturer",  "如：长春卡介苗研究所"),
    ("批签发号", "cert_no",       "如：2024S02345"),
    ("起始页",   "start_page",    "必填，大PDF中该证明首页页码（从1起）"),
    ("结束页",   "end_page",      "必填，大PDF中该证明末页页码（含末页）"),
    ("备注",     "notes",         "选填"),
]

def cmd_template(args: argparse.Namespace) -> None:
    out = Path(args.out)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "批签发索引"

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    hint_fill   = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    hdr_font    = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    hint_font   = Font(name="Calibri", italic=True, color="595959", size=9)
    center      = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # 第1行：列名
    for col, (label, _, _) in enumerate(TEMPLATE_COLS, 1):
        c = ws.cell(row=1, column=col, value=label)
        c.fill   = header_fill
        c.font   = hdr_font
        c.alignment = center

    # 第2行：填写说明
    for col, (_, _, hint) in enumerate(TEMPLATE_COLS, 1):
        c = ws.cell(row=2, column=col, value=hint)
        c.fill = hint_fill
        c.font = hint_font
        c.alignment = Alignment(wrap_text=True)

    # 示例数据行
    ws.append(["202501AB", "冻干人用狂犬病疫苗", "长春卡介苗研究所", "2025S00123", 1, 2, ""])
    ws.append(["202501CD", "冻干人用狂犬病疫苗", "长春卡介苗研究所", "2025S00124", 3, 5, "3页，已复核"])

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 10
    ws.column_dimensions["G"].width = 20
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 30

    wb.save(out)
    log.info(f"索引模板已生成: {out}")
    print(f"\n请用 Excel 打开 {out}，填写每个批号对应的起止页，保存后再运行 split。")
    print("  第1行：列名（勿删勿改）")
    print("  第2行：说明（可删）")
    print("  第3行起：每行一个批号\n")


# ── 读取 Excel 索引 ───────────────────────────────────────────
def _read_index_excel(path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    # 找列头行（第1行，跳过说明行）
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Excel 为空")

    header_map = {h: i for i, h in enumerate(rows[0]) if h}
    required = {"批号", "起始页", "结束页"}
    missing = required - set(header_map.keys())
    if missing:
        raise ValueError(f"索引表缺少必填列：{missing}")

    def _get(row, name):
        idx = header_map.get(name)
        return row[idx] if idx is not None else None

    _SKIP_VALUES = {"填写说明", "hint", "备注", "图例：", "图例"}

    records = []
    for row_num, row in enumerate(rows[1:], start=2):
        batch_no = str(_get(row, "批号") or "").strip()
        if not batch_no or batch_no.lower() in _SKIP_VALUES or batch_no.startswith("【"):
            continue  # 跳过空行、说明行、图例行、占位行

        raw_start = _get(row, "起始页")
        raw_end   = _get(row, "结束页")
        if raw_start is None or raw_end is None:
            log.debug(f"第 {row_num} 行跳过（起止页为空），批号={batch_no!r}")
            continue
        try:
            start = int(raw_start)
            end   = int(raw_end)
        except (TypeError, ValueError):
            raise ValueError(f"第 {row_num} 行：起始页/结束页必须为整数，批号={batch_no!r}")
        records.append({
            "batch_no":      batch_no,
            "vaccine_name":  str(_get(row, "疫苗名称") or "").strip() or None,
            "manufacturer":  str(_get(row, "生产企业") or "").strip() or None,
            "cert_no":       str(_get(row, "批签发号") or "").strip() or None,
            "start_page":    start,
            "end_page":      end,
            "notes":         str(_get(row, "备注") or "").strip() or None,
        })
    return records


# ── 质控 ─────────────────────────────────────────────────────
def _validate(records: list[dict], total_pages: int) -> list[str]:
    errors = []
    seen_ranges: list[tuple[int, int, str]] = []

    for i, r in enumerate(records, 1):
        s, e, bn = r["start_page"], r["end_page"], r["batch_no"]
        if s < 1:
            errors.append(f"行{i} 批号={bn}：起始页 {s} < 1")
        if e > total_pages:
            errors.append(f"行{i} 批号={bn}：结束页 {e} 超出 PDF 总页数 {total_pages}")
        if s > e:
            errors.append(f"行{i} 批号={bn}：起始页 {s} > 结束页 {e}")
        for ps, pe, pb in seen_ranges:
            if s <= pe and e >= ps:
                errors.append(f"行{i} 批号={bn} [{s}-{e}] 与 批号={pb} [{ps}-{pe}] 页码重叠")
        seen_ranges.append((s, e, bn))

    # 检查页码覆盖是否有空缺
    if records:
        all_pages = set()
        for r in records:
            all_pages.update(range(r["start_page"], r["end_page"] + 1))
        for pg in range(1, total_pages + 1):
            if pg not in all_pages:
                errors.append(f"警告：第 {pg} 页未被任何批号覆盖（可能遗漏）")

    return errors


# ── 自动识别核心 ──────────────────────────────────────────────
# 非国家批签发证明的企业报告关键词——命中则降为非首页
_COMPANY_REPORT_KW = re.compile(
    r'(?:产品质量报告|成品检验报告|成品检定报告|Certificate\s+of\s+Analysis'
    r'|药品生产许可证|产品检验报告)',
)


def _preprocess_text(text: str) -> str:
    """统一OCR常见误识：© → C（OCR将©字符误识为版权符号）"""
    return text.replace('©', 'C').replace('＀', '')


def _score_page(text: str) -> tuple[str, dict]:
    """
    对单页文字评分，返回 (置信度, 提取字段)。
    置信度: '高' | '中' | ''（非首页）
    """
    text = _preprocess_text(text)

    fields: dict[str, str | None] = {
        "batch_no": None, "vaccine_name": None,
        "manufacturer": None, "cert_no": None,
    }

    # ── 批号 ────────────────────────────────────────────────
    m = _RE_BATCH_NO.search(text)
    if m:
        raw = next((g for g in m.groups() if g), "").strip()
        raw = raw.rstrip('.')   # 去除 OCR 误产生的尾部句点：202409009. → 202409009
        fields["batch_no"] = raw or None

    # ── 批签发号（优先国际格式，保留完整前缀） ──────────────────
    m = _RE_CERT_NO_INTL.search(text)
    if m:
        fields["cert_no"] = m.group(0).strip()
    else:
        m = _RE_CERT_NO_CN.search(text)
        if m:
            fields["cert_no"] = m.group(0).strip()

    # ── 生产企业 ─────────────────────────────────────────────
    m = _RE_MANUFACTURER_NEXT.search(text)
    if not m:
        m = _RE_MANUFACTURER.search(text)
    if m:
        fields["manufacturer"] = m.group(1).strip() or None

    # ── 疫苗名称 ─────────────────────────────────────────────
    m = _RE_VACCINE_GENERIC.search(text)
    if m:
        fields["vaccine_name"] = m.group(1).strip() or None
    else:
        candidates = _RE_VACCINE_LINE.findall(text)
        if candidates:
            fields["vaccine_name"] = min(candidates, key=len).strip() or None

    # ── 评分 ────────────────────────────────────────────────
    # 高置信优先：含国家批签发证明确切标题，不受企业报告过滤影响
    for kw in _HIGH_KW_EXACT:
        if kw in text:
            return "高", fields
    if _HIGH_KW_LINE.search(text):
        return "高", fields

    # 企业内部报告页（检验报告/生产许可证等）排除在外，不纳入中置信
    if _COMPANY_REPORT_KW.search(text):
        return "", {}

    # 中置信：含批签发编号上下文 + 批号 + 辅助字段（三者同时满足）
    has_release_ctx = bool(re.search(r'\bLR[A-Z]\d{5}', text))  # LRA/LRH/LRG...
    has_batch       = bool(fields["batch_no"])
    has_aux         = any(kw in text for kw in ("生产企业", "有效期至", "Manufacturer"))
    if has_release_ctx and has_batch and has_aux:
        return "中", fields

    return "", {}


def _build_detect_index(
    detections: list[dict], total_pages: int
) -> tuple[list[dict], list[str]]:
    """
    将检测到的疑似首页列表转换为起止页索引，并生成警告信息。
    detections: [{"page": int, "conf": str, "source": str, "fields": dict}, ...]
    返回: (records, warnings)
    """
    warnings: list[str] = []
    records: list[dict] = []

    if not detections:
        warnings.append("未检测到任何疑似首页，请完全手动填写复核表。")
        return records, warnings

    # 第 1 页未被识别为首页：说明有前置页面未覆盖
    if detections[0]["page"] > 1:
        warnings.append(
            f"第 1-{detections[0]['page'] - 1} 页未被识别为任何证明首页，"
            "已在复核表顶部添加【待核查】占位行，请手动确认。"
        )
        records.append({
            "batch_no":     "【待核查】",
            "vaccine_name": None,
            "manufacturer": None,
            "cert_no":      None,
            "start_page":   1,
            "end_page":     detections[0]["page"] - 1,
            "conf":         "手动",
            "source":       "manual",
            "notes":        "自动检测未覆盖，请核查",
        })

    for i, det in enumerate(detections):
        end = detections[i + 1]["page"] - 1 if i + 1 < len(detections) else total_pages
        f   = det["fields"]
        records.append({
            "batch_no":     f.get("batch_no") or f"【第{det['page']}页】",
            "vaccine_name": f.get("vaccine_name"),
            "manufacturer": f.get("manufacturer"),
            "cert_no":      f.get("cert_no"),
            "start_page":   det["page"],
            "end_page":     end,
            "conf":         det["conf"],
            "source":       det.get("source", "ocr"),
            "notes":        "自动识别" if det["conf"] == "高" else "中置信，请核查",
        })

    return records, warnings


def _save_detect_excel(records: list[dict], out_path: Path) -> None:
    """将检测结果保存为带颜色标注的复核 Excel。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "批签发复核表"

    # 颜色
    fill_header = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    fill_high   = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fill_mid    = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    fill_manual = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    font_hdr    = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    center      = Alignment(horizontal="center", vertical="center", wrap_text=True)

    headers = ["批号", "疫苗名称", "生产企业", "批签发号", "起始页", "结束页", "置信度", "备注"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill      = fill_header
        c.font      = font_hdr
        c.alignment = center

    _CONF_FILL = {"高": fill_high, "中": fill_mid, "手动": fill_manual}

    for row_idx, r in enumerate(records, 2):
        row_data = [
            r["batch_no"], r["vaccine_name"], r["manufacturer"], r["cert_no"],
            r["start_page"], r["end_page"], r["conf"], r["notes"],
        ]
        fill = _CONF_FILL.get(r["conf"], fill_manual)
        for col, val in enumerate(row_data, 1):
            c = ws.cell(row=row_idx, column=col, value=val)
            c.fill      = fill
            c.alignment = Alignment(vertical="center", wrap_text=True)

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 10
    ws.column_dimensions["G"].width = 10
    ws.column_dimensions["H"].width = 22
    ws.row_dimensions[1].height = 22

    # 图例说明行（最后）
    legend_row = len(records) + 3
    ws.cell(row=legend_row, column=1,
            value="图例：").font = Font(bold=True)
    for col, (label, fill) in enumerate([
        ("高置信（自动识别可靠）", fill_high),
        ("中置信（建议核查）", fill_mid),
        ("手动（需人工填写）", fill_manual),
    ], 2):
        c = ws.cell(row=legend_row, column=col, value=label)
        c.fill = fill

    wb.save(out_path)


def cmd_detect(args: argparse.Namespace) -> None:
    """v2 自动识别疑似首页，生成人工复核表。"""
    pdf_path = Path(args.pdf).resolve()
    if not pdf_path.exists():
        log.error(f"PDF 文件不存在: {pdf_path}")
        sys.exit(1)

    out_path = Path(args.out) if args.out else pdf_path.parent / f"{pdf_path.stem}_复核表.xlsx"
    use_ocr  = getattr(args, "ocr", False)
    dpi      = getattr(args, "dpi", 200)

    log.info(f"打开 PDF: {pdf_path}")
    src = fitz.open(str(pdf_path))
    total_pages = len(src)
    log.info(f"PDF 总页数: {total_pages}")

    engine = "macOS Vision" if sys.platform == "darwin" else "Tesseract"
    use_qr = not getattr(args, "no_qr", False)   # 默认开启二维码模式
    if use_qr:
        try:
            import cv2  # noqa: F401
            log.info("二维码模式已启用（需联网调用国家药监局 API）")
        except ImportError:
            log.warning("opencv-python 未安装，二维码模式不可用，将仅使用 OCR")
            use_qr = False

    if use_ocr:
        log.info(f"OCR 模式（{engine}，DPI={dpi}）——扫描版 PDF")
    else:
        # 自动检测：若第1页文字量极少则切换 OCR
        sample = src[0].get_text().strip()
        if len(sample) < 10:
            log.warning("第1页提取文字过少（可能为扫描件），自动启用 OCR 模式")
            use_ocr = True
        else:
            log.info("文字提取模式——数字 PDF")

    # ── 透视矫正（仅扫描件模式） ─────────────────────────────
    # 预先检测文档边界并矫正，改善 QR 解码率和 OCR 准确率
    use_dewarp = use_ocr  # 仅扫描件才需要矫正；数字 PDF 跳过
    try:
        import numpy as _np
        import cv2 as _cv2
        _dewarp_available = True
    except ImportError:
        _dewarp_available = False
        use_dewarp = False

    def _get_dewarped(page: "fitz.Page") -> "np.ndarray | None":
        """以 scale=3 渲染当前页并尝试透视矫正，返回矫正后 RGB 图像或 None。"""
        if not _dewarp_available:
            return None
        try:
            pix = page.get_pixmap(matrix=fitz.Matrix(3, 3))
            raw = _np.frombuffer(pix.samples, dtype=_np.uint8).reshape(
                pix.height, pix.width, pix.n
            )
            rgb = _cv2.cvtColor(raw, _cv2.COLOR_RGBA2RGB) if pix.n == 4 else raw
            return _dewarp_image(rgb)
        except Exception:
            return None

    # ── 逐页扫描 ────────────────────────────────────────────
    detections: list[dict] = []
    qr_hits = 0
    log.info("开始扫描各页……")
    for i in range(total_pages):
        page = src[i]

        # ── 透视矫正（扫描件） ──────────────────────────────
        dewarped = _get_dewarped(page) if use_dewarp else None
        if dewarped is not None:
            log.debug(f"  第{i+1}页 透视矫正 OK ({dewarped.shape[1]}x{dewarped.shape[0]})")

        # ── 优先尝试二维码 ──────────────────────────────────
        if use_qr:
            qr_content = _decode_qr_page(page)
            if qr_content and "nmpa.gov.cn" in qr_content:
                fields = _fetch_nmpa_cert(qr_content)
                if fields and fields.get("batch_no"):
                    qr_hits += 1
                    log.info(
                        f"  QR ✓ 第 {i+1} 页  批号={fields['batch_no']}  "
                        f"批签发号={fields.get('cert_no') or '—'}"
                    )
                    detections.append({"page": i + 1, "conf": "高", "source": "qr", "fields": fields})
                    continue   # 不再做 OCR

        # ── 回退到 OCR / 文字提取 ──────────────────────────
        if use_ocr:
            try:
                # 优先使用矫正后图像（OCR 准确率更高）
                ocr_img = None
                if dewarped is not None:
                    import numpy as np
                    scale = dpi / 72 / 3  # 已在 scale=3 下矫正，等比换算到目标 DPI
                    if scale > 1.0:
                        ocr_img = _cv2.resize(
                            dewarped,
                            (int(dewarped.shape[1] * scale), int(dewarped.shape[0] * scale)),
                            interpolation=_cv2.INTER_CUBIC,
                        )
                    else:
                        ocr_img = dewarped
                text = _ocr_page(page, dpi=dpi, img=ocr_img)
            except RuntimeError as e:
                log.error(str(e))
                sys.exit(1)
        else:
            text = page.get_text()

        conf, fields = _score_page(text)
        status = f"[{conf}]" if conf else "  -  "
        log.debug(f"  第{i+1:>3}页  {status}  {text[:40].replace(chr(10),' ')}")
        if conf:
            log.info(f"  OCR 第 {i+1} 页  置信度={conf}  批号={fields.get('batch_no') or '未提取'}")
            detections.append({"page": i + 1, "conf": conf, "source": "ocr", "fields": fields})

    # ── 构建索引 ────────────────────────────────────────────
    records, warnings = _build_detect_index(detections, total_pages)
    for w in warnings:
        log.warning(w)

    # ── 保存复核表 ───────────────────────────────────────────
    _save_detect_excel(records, out_path)

    # ── 控制台摘要 ───────────────────────────────────────────
    high   = sum(1 for r in records if r["conf"] == "高")
    mid    = sum(1 for r in records if r["conf"] == "中")
    manual = sum(1 for r in records if r["conf"] == "手动")

    sep = "─" * 60
    print(f"\n{sep}")
    print(f"  自动识别报告")
    print(sep)
    print(f"  PDF 总页数  : {total_pages}")
    print(f"  检测到证明  : {len(records)} 份")
    if use_qr:
        print(f"  二维码命中  : {qr_hits}  （字段来自国家药监局）")
    print(f"  高置信 🟢   : {high}  （自动识别可靠）")
    print(f"  中置信 🟡   : {mid}  （建议人工核查）")
    print(f"  手动   🔴   : {manual}  （未检测，需手工填写）")
    print(sep)
    print(f"  {'批号':<22} {'页':<8} {'页数':>4}  置信度")
    print(f"  {'─'*22} {'─'*8} {'─'*4}  ───")
    for r in records:
        icon = {"高": "🟢", "中": "🟡", "手动": "🔴"}.get(r["conf"], "")
        page_str = f"{r['start_page']}-{r['end_page']}"
        pg = r["end_page"] - r["start_page"] + 1
        print(f"  {r['batch_no']:<22} {page_str:<8} {pg:>4}  {icon} {r['conf']}")
    print(sep)
    print(f"\n复核表已保存: {out_path}")
    print("请用 Excel 打开，核对后直接运行：")
    print(f"  python3 tools/batch_cert_splitter.py split --pdf {pdf_path.name} --index {out_path.name}\n")


# ── 核心拆分 ─────────────────────────────────────────────────
def cmd_split(args: argparse.Namespace) -> None:
    pdf_path   = Path(args.pdf).resolve()
    index_path = Path(args.index).resolve()
    out_dir    = Path(args.out_dir).resolve() if args.out_dir else OUTPUT_DIR / pdf_path.stem

    if not pdf_path.exists():
        log.error(f"PDF 文件不存在: {pdf_path}")
        sys.exit(1)
    if not index_path.exists():
        log.error(f"索引文件不存在: {index_path}")
        sys.exit(1)

    out_dir.mkdir(parents=True, exist_ok=True)
    init_db()

    # ── 读索引 ──────────────────────────────────────────────
    log.info(f"读取索引: {index_path}")
    records = _read_index_excel(index_path)
    log.info(f"共 {len(records)} 条记录")

    # ── 打开 PDF ────────────────────────────────────────────
    log.info(f"打开 PDF: {pdf_path}")
    src = fitz.open(str(pdf_path))
    total_pages = len(src)
    log.info(f"PDF 总页数: {total_pages}")

    # ── 质控 ────────────────────────────────────────────────
    errors = _validate(records, total_pages)
    hard_errors = [e for e in errors if not e.startswith("警告")]
    for e in errors:
        if e.startswith("警告"):
            log.warning(e)
        else:
            log.error(e)
    if hard_errors:
        log.error(f"发现 {len(hard_errors)} 个错误，终止拆分。请修正索引表后重试。")
        sys.exit(1)

    # ── 注册源文件 ──────────────────────────────────────────
    now = datetime.now().isoformat(timespec="seconds")
    with get_conn() as conn:
        cur = conn.execute(
            "INSERT OR IGNORE INTO batch_files "
            "(original_filename, original_path, total_pages, import_time, status) "
            "VALUES (?,?,?,?,'pending')",
            (pdf_path.name, str(pdf_path), total_pages, now)
        )
        if cur.lastrowid:
            file_id = cur.lastrowid
        else:
            file_id = conn.execute(
                "SELECT id FROM batch_files WHERE original_path=?", (str(pdf_path),)
            ).fetchone()["id"]

    # ── 逐条拆分 ────────────────────────────────────────────
    results = []
    for r in records:
        s, e = r["start_page"], r["end_page"]   # 1-indexed
        page_count = e - s + 1

        # 文件命名：批号_批签发号.pdf 或 批号.pdf
        stem = r["batch_no"]
        if r["cert_no"]:
            stem = f"{r['batch_no']}_{r['cert_no']}"
        out_pdf = out_dir / f"{stem}.pdf"

        # 复核状态：页数异常则标为 abnormal
        review = "ok"
        notes_parts = [r["notes"]] if r["notes"] else []
        if not (NORMAL_PAGE_RANGE[0] <= page_count <= NORMAL_PAGE_RANGE[1]):
            review = "abnormal"
            notes_parts.append(f"页数={page_count}，超出正常范围{NORMAL_PAGE_RANGE}，需人工复核")

        # 用 PyMuPDF 提取页面（0-indexed）
        writer = fitz.open()
        writer.insert_pdf(src, from_page=s - 1, to_page=e - 1)
        writer.save(str(out_pdf))
        writer.close()
        log.info(f"  拆分: {r['batch_no']} [{s}-{e}页] → {out_pdf.name}  {review}")

        results.append({**r,
                        "page_count":   page_count,
                        "output_pdf":   str(out_pdf),
                        "review_status": review,
                        "notes":        "；".join(notes_parts) or None})

    # ── 写入数据库 ───────────────────────────────────────────
    split_time = datetime.now().isoformat(timespec="seconds")
    with get_conn() as conn:
        for r in results:
            conn.execute(
                "INSERT INTO cert_index "
                "(batch_file_id, batch_no, vaccine_name, manufacturer, cert_no, "
                " start_page, end_page, page_count, output_pdf, review_status, split_time, notes) "
                "VALUES (?,?,?,?,?,?,?,?,?,?,?,?)",
                (file_id, r["batch_no"], r["vaccine_name"], r["manufacturer"],
                 r["cert_no"], r["start_page"], r["end_page"], r["page_count"],
                 r["output_pdf"], r["review_status"], split_time, r["notes"])
            )
        conn.execute("UPDATE batch_files SET status='split' WHERE id=?", (file_id,))

    # ── 拆分报告 ────────────────────────────────────────────
    _print_report(pdf_path, total_pages, results, out_dir)


def _print_report(pdf_path: Path, total_pages: int, results: list[dict], out_dir: Path) -> None:
    ok_count       = sum(1 for r in results if r["review_status"] == "ok")
    abnormal_count = sum(1 for r in results if r["review_status"] == "abnormal")

    sep = "─" * 60
    print(f"\n{sep}")
    print(f"  拆分报告")
    print(sep)
    print(f"  源文件     : {pdf_path.name}")
    print(f"  PDF总页数  : {total_pages}")
    print(f"  拆分总数   : {len(results)}")
    print(f"  正常       : {ok_count}")
    print(f"  待复核     : {abnormal_count}")
    print(f"  输出目录   : {out_dir}")
    print(sep)
    print(f"  {'批号':<20} {'页':<8} {'页数':>4}  状态")
    print(f"  {'─'*20} {'─'*8} {'─'*4}  ──────")
    for r in results:
        flag = "⚠ 待复核" if r["review_status"] == "abnormal" else "✓"
        print(f"  {r['batch_no']:<20} {r['start_page']}-{r['end_page']:<5} {r['page_count']:>4}  {flag}")
    print(sep)
    if abnormal_count:
        print("  ⚠ 以下批号页数异常，请人工核查：")
        for r in results:
            if r["review_status"] == "abnormal":
                print(f"    - {r['batch_no']}  {r['start_page']}-{r['end_page']}页  ({r['notes']})")
    print()


# ── list 命令 ────────────────────────────────────────────────
def cmd_list(args: argparse.Namespace) -> None:
    init_db()
    limit = args.limit
    with get_conn() as conn:
        rows = conn.execute(
            "SELECT c.batch_no, c.vaccine_name, c.cert_no, c.start_page, c.end_page, "
            "       c.page_count, c.review_status, b.original_filename "
            "FROM cert_index c JOIN batch_files b ON b.id=c.batch_file_id "
            "ORDER BY c.id DESC LIMIT ?",
            (limit,)
        ).fetchall()
    if not rows:
        print("数据库中暂无记录。")
        return
    print(f"\n最近 {len(rows)} 条记录：")
    print(f"  {'批号':<20} {'疫苗名称':<20} {'批签发号':<15} {'页':<8} {'页数':>4}  状态       源文件")
    print("  " + "─" * 95)
    for r in rows:
        print(f"  {r['batch_no']:<20} {(r['vaccine_name'] or ''):<20} "
              f"{(r['cert_no'] or ''):<15} "
              f"{r['start_page']}-{r['end_page']:<4} {r['page_count']:>4}  "
              f"{r['review_status']:<10} {r['original_filename']}")
    print()


# ── CLI ──────────────────────────────────────────────────────
def main() -> None:
    parser = argparse.ArgumentParser(
        description="批签发大PDF拆分工具 v2",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    sub = parser.add_subparsers(dest="cmd", required=True)

    sub.add_parser("init", help="初始化 SQLite 数据库")

    tp = sub.add_parser("template", help="生成 Excel 索引模板（v1 手动流程）")
    tp.add_argument("--out", default="批签发索引模板.xlsx", help="输出文件名")

    dp = sub.add_parser("detect", help="自动识别疑似首页，生成人工复核表（v2 流程）")
    dp.add_argument("--pdf", required=True, help="大 PDF 路径")
    dp.add_argument("--out", default=None,  help="复核表输出路径（默认 <pdf名>_复核表.xlsx）")
    dp.add_argument("--ocr", action="store_true",
                    help="强制启用 OCR（扫描版 PDF 必须加此参数，或程序自动检测后启用）")
    dp.add_argument("--dpi", type=int, default=200,
                    help="OCR 渲染分辨率，默认 200；扫描不清晰时可提至 300")
    dp.add_argument("--no-qr", action="store_true",
                    help="禁用二维码模式（离线环境或不需要联网时使用）")

    sp = sub.add_parser("split", help="按索引/复核表拆分 PDF，写入数据库")
    sp.add_argument("--pdf",     required=True, help="大 PDF 路径")
    sp.add_argument("--index",   required=True, help="已填写/复核的 Excel 表路径")
    sp.add_argument("--out-dir", default=None,  help="拆分 PDF 输出目录（默认 output/<pdf名>/）")

    lp = sub.add_parser("list", help="查看已入库的单证记录")
    lp.add_argument("--limit", type=int, default=50, help="最多显示条数（默认50）")

    args = parser.parse_args()

    if args.cmd == "init":
        init_db()
    elif args.cmd == "template":
        cmd_template(args)
    elif args.cmd == "detect":
        cmd_detect(args)
    elif args.cmd == "split":
        cmd_split(args)
    elif args.cmd == "list":
        cmd_list(args)


if __name__ == "__main__":
    main()
